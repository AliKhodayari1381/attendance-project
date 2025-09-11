from rest_framework import viewsets
from .models import Employee, Attendance,WorkSchedule
from .serializers import EmployeeSerializer, AttendanceSerializer
from rest_framework.decorators import api_view, permission_classes,action
from rest_framework.permissions import IsAuthenticated,AllowAny,IsAdminUser
from rest_framework.response import Response
from datetime import datetime
import jdatetime
from openpyxl import load_workbook
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from copy import copy

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAdminUser,IsAuthenticated]

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status','employee']
    http_method_names = ['get', 'patch', 'delete']
    permission_classes = [IsAdminUser,IsAuthenticated]

    def get_queryset(self):
        
        queryset = Attendance.objects.all()
        date_param = self.request.query_params.get('date')

        if date_param:  
            try:
                parts = [int(p) for p in date_param.split('-')]
                date_obj = jdatetime.date(*parts)
                queryset = queryset.filter(date=date_obj)
            except Exception:
                date_obj = jdatetime.date.today()
        else:
            date_obj = jdatetime.date.today()

        queryset = queryset.filter(date=date_obj)
        
        return queryset


    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        خروجی اکسل بر اساس تاریخ انتخاب شده (یا امروز)
        """

        date_param = request.query_params.get('date')
        if date_param:
            try:
                parts = [int(p) for p in date_param.split('-')]
                date_obj = jdatetime.date(*parts)
            except Exception:
                date_obj = jdatetime.date.today()
        else:
            date_obj = jdatetime.date.today()

        queryset = Attendance.objects.filter(date=date_obj)

        wb = load_workbook('attendance_template.xlsx')
        ws = wb.active

        template_row = 2  
        start_row = 2

        for row_num, attendance in enumerate(queryset, start=start_row):
            ws[f"A{row_num}"].value = attendance.id
            ws[f"B{row_num}"].value = str(attendance.employee) if attendance.employee else ''
            ws[f"C{row_num}"].value = attendance.date.strftime('%Y-%m-%d')
            ws[f"D{row_num}"].value = attendance.check_in_time.strftime('%H:%M:%S') if attendance.check_in_time else ''
            ws[f"E{row_num}"].value = attendance.get_status_display() if attendance.status else ''
            ws[f"F{row_num}"].value = f"{attendance.late_minutes} دقیقه" if attendance.late_minutes is not None else ''

            for col_num in range(1, 7): 
                source_cell = ws.cell(row=template_row, column=col_num)
                target_cell = ws.cell(row=row_num, column=col_num)
                if row_num != template_row:  
                    target_cell._style = copy(source_cell._style)

        filename = f"{date_obj.strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response



@api_view(['POST'])
@permission_classes([AllowAny])
def check_in(request):
    """
    ثبت حضور کاربر
    """
    employee_id = request.data.get('employee_id')
    try:
        employee = Employee.objects.get(employee_id=employee_id)
    except Employee.DoesNotExist:
        return Response({"خطا": "کاربر یافت نشد"}, status=404)

    # بررسی اگر حضور امروز ثبت شده
    today = datetime.today().date()
    if Attendance.objects.filter(employee=employee, date=today).exists():
        return Response({"پیغام": "این کاربر قبلا ثبت کرده است"}, status=400)

    now = datetime.now()
    schedule = WorkSchedule.objects.first()
    work_start = schedule.start_time

    status=Attendance.Status.PRESENT.value
    late_minutes = 0

    if now.time() > work_start:
        status=Attendance.Status.LATE.value

        diff = datetime.combine(today, now.time()) - datetime.combine(today, work_start)
        late_minutes = diff.seconds // 60

    attendance = Attendance.objects.create(
        employee = employee , 
        date= today,
        check_in_time = now.time(),
        status = status,
        late_minutes = late_minutes

    )
    jalali_date = attendance.date.strftime("%Y-%m-%d")
    return Response({
        "پیغام": "با موفقیت ثبت شد",
        "کارمند": employee.first_name +' '+employee.last_name,
        "تاریخ": jalali_date,
        "زمان": attendance.check_in_time.strftime("%H:%M"),
        "وضعیت": attendance.get_status_display(),
        "مدت زمان تاخیر": f"{attendance.late_minutes} دقیقه"

    })
    